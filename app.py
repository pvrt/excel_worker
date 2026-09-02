from collections import defaultdict
import hashlib
import os
from pathlib import Path
import platform
import re
import shutil
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl


# ==============================================================================
#                             ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==============================================================================


def clean_val(val) -> str:
    """Очищает значение ячейки."""
    if val is None:
        return ""
    s = str(val).replace("\xa0", " ").strip()
    if isinstance(val, float) and val.is_integer():
        s = str(int(val))
    elif s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def extract_last_number_from_filename(filename: str) -> str:
    """Извлекает последнюю группу цифр из имени файла."""
    stem = Path(filename).stem
    numbers = re.findall(r"\d+", stem)
    return numbers[-1] if numbers else stem


def get_file_based_result(
    filepath: Path, base_dir: Path, result_type: str
) -> str:
    if result_type == "last_number":
        return extract_last_number_from_filename(filepath.name)
    elif result_type == "full_filename":
        return filepath.name
    elif result_type == "filename_stem":
        return filepath.stem
    elif result_type == "relative_path":
        try:
            return str(filepath.relative_to(base_dir))
        except ValueError:
            return filepath.name
    return filepath.name


def get_fixed_cell_value(filepath: Path, row_num: int, col_num: int) -> str:
    """Считывает значение из конкретной строки и колонки первого листа файла."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        for _, row in enumerate(
            sheet.iter_rows(
                min_row=row_num, max_row=row_num, values_only=True
            ),
            start=row_num,
        ):
            if len(row) >= col_num and row[col_num - 1] is not None:
                wb.close()
                return clean_val(row[col_num - 1])
        wb.close()
    except Exception:
        pass
    return ""


def calculate_file_hash(filepath: Path, algorithm: str = "sha256") -> str:
    """Вычисляет хеш файла по частям (блоками по 64 КБ)."""
    hasher = getattr(hashlib, algorithm.lower())()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def sanitize_filename(name: str) -> str:
    """Удаляет недопустимые для имени файла символы и обрезает пробелы."""
    # Заменяем недопустимые символы Windows: \ / : * ? " < > |
    sanitized = re.sub(r'[\\/:*?"<>|]', "_", name).strip()
    # Убираем точки/пробелы в конце (Windows не любит)
    sanitized = sanitized.strip(" .")
    # Ограничиваем длину
    if len(sanitized) > 180:
        sanitized = sanitized[:180]
    return sanitized or "document"


def find_soffice_executable() -> str | None:
    """Ищет исполняемый файл LibreOffice (soffice)."""
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    # Попытка через PATH (Linux/macOS/портативная установка)
    for name in ("soffice", "soffice.exe", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    return None


def _is_excel_available() -> bool:
    """Проверяет, доступен ли Microsoft Excel через COM (только Windows)."""
    if platform.system() != "Windows":
        return False
    # Проверка через реестр
    try:
        import winreg

        try:
            key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application")
            winreg.CloseKey(key)
            return True
        except FileNotFoundError:
            pass
        try:
            key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"CLSID\{00024500-0000-0000-C000-000000000046}")
            winreg.CloseKey(key)
            return True
        except FileNotFoundError:
            pass
    except Exception:
        pass
    # Проверка наличия EXCEL.EXE в стандартных путях
    excel_paths = [
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\root\Office15\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office15\EXCEL.EXE",
    ]
    for p in excel_paths:
        if os.path.isfile(p):
            return True
    # Последняя попытка — где лежит excel через PATH/which (редко)
    if shutil.which("excel") or shutil.which("EXCEL.EXE"):
        return True
    return False


GOOGLE_DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]


def _get_google_drive_service(credentials_path: str = "credentials.json", token_path: str = "token.json"):
    """Создаёт сервис Google Drive с OAuth (credentials.json -> token.json)."""
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
        from google.auth.transport.requests import Request
    except ImportError as e:
        raise RuntimeError(
            "Не установлены Google библиотеки. Выполните:\n"
            "pip install google-api-python-client google-auth google-auth-oauthlib requests"
        ) from e

    creds = None
    if os.path.exists(token_path):
        try:
            creds = Credentials.from_authorized_user_file(token_path, GOOGLE_DRIVE_SCOPES)
        except Exception:
            creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds or not creds.valid:
            if not os.path.exists(credentials_path):
                raise FileNotFoundError(
                    f"Не найден файл {credentials_path}.\n"
                    "Скачайте его в Google Cloud Console -> APIs & Services -> Credentials -> Create OAuth client ID (Desktop)."
                )
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, GOOGLE_DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
            # Сохраняем токен
            try:
                with open(token_path, "w", encoding="utf-8") as f:
                    f.write(creds.to_json())
            except Exception:
                pass
    try:
        service = build("drive", "v3", credentials=creds)
    except Exception as e:
        raise RuntimeError(f"Не удалось создать Google Drive сервис: {e}") from e
    return service, creds


def _convert_single_xlsx_to_pdf_google(drive_service, creds, xlsx_path: Path, pdf_path: Path) -> None:
    """Конвертирует один xlsx в pdf через Google Drive/Sheets (загрузка -> экспорт -> удаление)."""
    try:
        from googleapiclient.http import MediaFileUpload
        import requests
    except ImportError as e:
        raise RuntimeError("Не установлены зависимости Google. pip install google-api-python-client requests") from e

    # Обновляем токен если истёк
    try:
        if creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request

            creds.refresh(Request())
    except Exception:
        pass

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    file_metadata = {
        "name": xlsx_path.stem,
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    uploaded = drive_service.files().create(body=file_metadata, media_body=media, fields="id").execute()
    file_id = uploaded.get("id")
    if not file_id:
        raise RuntimeError("Google Drive не вернул fileId")
    try:
        # Параметры экспорта как в примере: A4, альбом, fitw, сетка, повтор шапки
        export_url = (
            f"https://docs.google.com/spreadsheets/d/{file_id}/export?"
            f"exportFormat=pdf&format=pdf"
            f"&size=A4"
            f"&portrait=false"
            f"&fitw=true"
            f"&gridlines=true"
            f"&fzr=true"
        )
        headers = {"Authorization": f"Bearer {creds.token}"}
        resp = requests.get(export_url, headers=headers, timeout=120)
        resp.raise_for_status()
        with open(pdf_path, "wb") as f:
            f.write(resp.content)
        if pdf_path.stat().st_size == 0:
            raise RuntimeError("Google вернул пустой PDF")
    finally:
        try:
            drive_service.files().delete(fileId=file_id).execute()
        except Exception:
            pass


def _convert_single_xlsx_to_pdf_excel(excel_app, xlsx_path: Path, pdf_path: Path) -> None:
    """Конвертирует один xlsx в pdf через уже запущенный Excel COM-объект. Стараемся убрать лишние пустые страницы."""
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    abs_src = str(xlsx_path.resolve())
    abs_dst = str(pdf_path.resolve())
    wb = excel_app.Workbooks.Open(abs_src)
    try:
        # Пытаемся ограничить печать только используемым диапазоном и подогнать по ширине
        try:
            ws = wb.ActiveSheet
            used = ws.UsedRange
            # Если UsedRange не пустой — задаём область печати
            if used is not None:
                try:
                    ws.PageSetup.PrintArea = used.Address
                except Exception:
                    pass
                try:
                    ws.PageSetup.Zoom = False
                    ws.PageSetup.FitToPagesWide = 1
                    ws.PageSetup.FitToPagesTall = 0  # в высоту — сколько нужно, без пустых страниц
                    ws.PageSetup.Orientation = 2 if used.Columns.Count > 6 else 1  # 2=landscape
                except Exception:
                    pass
                try:
                    # Убираем лишние поля
                    ws.PageSetup.LeftMargin = excel_app.InchesToPoints(0.25)
                    ws.PageSetup.RightMargin = excel_app.InchesToPoints(0.25)
                    ws.PageSetup.TopMargin = excel_app.InchesToPoints(0.3)
                    ws.PageSetup.BottomMargin = excel_app.InchesToPoints(0.3)
                except Exception:
                    pass
        except Exception:
            pass
        # 0 = xlTypePDF
        wb.ExportAsFixedFormat(0, abs_dst)
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass


def _convert_single_xlsx_to_pdf_libreoffice(
    soffice: str, xlsx_path: Path, out_dir: Path, timeout: int = 90
) -> Path | None:
    """Конвертирует один xlsx в pdf через LibreOffice, возвращает путь к созданному pdf.

    Использует временную папку чтобы избежать коллизий при одинаковых stem.
    Перед конвертацией создаёт временную копию с подрезанной областью печати,
    чтобы убрать 2-3x пустого места после таблицы (часто из-за форматирования далёких ячеек).
    """
    import tempfile
    import time
    from openpyxl.utils import get_column_letter

    # Создаём временную папку для промежуточного pdf
    tmpdir = Path(tempfile.mkdtemp(prefix="xlsx2pdf_"))
    # Подготавливаем копию с корректной областью печати
    prep_path = tmpdir / xlsx_path.name
    src_for_convert = xlsx_path
    try:
        shutil.copy2(str(xlsx_path), str(prep_path))
        try:
            wb = openpyxl.load_workbook(prep_path)
            for ws in wb.worksheets:
                # Находим последнюю непустую строку/колонку по значениям (как для reportlab)
                last_row = 0
                last_col = 0
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    has = False
                    c_idx_last = 0
                    for c_idx, v in enumerate(row, 1):
                        if v is not None and str(v).strip() != "":
                            has = True
                            c_idx_last = c_idx
                    if has:
                        last_row = r_idx
                        if c_idx_last > last_col:
                            last_col = c_idx_last
                # Учитываем merged cells
                try:
                    for rng in ws.merged_cells.ranges:
                        if rng.max_row > last_row:
                            last_row = rng.max_row
                        if rng.max_col > last_col:
                            last_col = rng.max_col
                except Exception:
                    pass
                if last_row > 0 and last_col > 0:
                    # Ограничиваем, чтобы не резать заголовки/подписи внизу (если они есть)
                    # но убираем далёкий форматированный хвост (4705 -> реально 4705 и есть данные, так что оставим)
                    # Ставим область печати A1:Last
                    try:
                        ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
                    except Exception:
                        pass
                    # Настраиваем страницу, чтобы в PDF не было 2-3x пустого хвоста и страниц
                    try:
                        # Сохраняем оригинальный масштаб/фит, если уже задан (например scale=54 у файла 375)
                        has_scale = ws.page_setup.scale not in (None, 0)
                        has_fit = ws.page_setup.fitToWidth not in (None, 0) or ws.page_setup.fitToHeight not in (None, 0)
                        if not has_scale and not has_fit:
                            ws.page_setup.fitToWidth = 1
                            ws.page_setup.fitToHeight = 0
                            ws.sheet_properties.pageSetUpPr.fitToPage = True
                        if not ws.page_setup.orientation:
                            ws.page_setup.orientation = "landscape" if last_col > 6 else "portrait"
                        if not ws.page_setup.paperSize:
                            ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    except Exception:
                        pass
            wb.save(prep_path)
            wb.close()
            src_for_convert = prep_path
        except Exception:
            # Если не удалось подготовить — используем оригинал
            src_for_convert = xlsx_path
    except Exception:
        src_for_convert = xlsx_path

    try:
        expected = tmpdir / f"{xlsx_path.stem}.pdf"
        # Используем calc_pdf_Export с IsSkipEmptyPages, чтобы не плодить пустые страницы/листы
        # Для .xls/.xlsx это Calc; для других типов LibreOffice сам выберет фильтр, но указание calc безопасно
        convert_filter = "pdf:calc_pdf_Export:IsSkipEmptyPages=true;SelectPdfVersion=1"
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", convert_filter, "--outdir", str(tmpdir), str(src_for_convert)],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        # Fallback без опций, если фильтр не поддерживается старой версией
        if result.returncode != 0 and "Error" in (result.stderr or ""):
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmpdir), str(src_for_convert)],
                capture_output=True,
                text=True,
                timeout=timeout,
            )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip() or f"soffice exit {result.returncode}")
        if not expected.exists():
            for _ in range(5):
                time.sleep(0.3)
                if expected.exists():
                    break
            if not expected.exists():
                # Пробуем найти любой pdf в tmpdir
                pdfs = list(tmpdir.glob("*.pdf"))
                if pdfs:
                    expected = pdfs[0]
                else:
                    raise FileNotFoundError(f"LibreOffice не создал файл: {expected}")
        # Перемещаем во временный путь, возвращаем путь к tmp pdf для последующего переименования
        # Копируем в out_dir как временный файл, вызывающий код сам переименует
        # Возвращаем путь к промежуточному файлу
        return expected
    except Exception:
        # Очистка при ошибке
        try:
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass
        raise


def _convert_single_xlsx_to_pdf_reportlab(xlsx_path: Path, pdf_path: Path) -> None:
    """Fallback конвертация через openpyxl + reportlab (табличный вид без точного форматирования)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_LEFT
    except ImportError as e:
        raise RuntimeError("Не установлен reportlab. Установите: pip install reportlab") from e

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        # Пропускаем полностью пустые строки
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append([clean_val(v) for v in row])

    wb.close()

    if not rows:
        rows = [["(пустой лист)"]]

    # Ограничиваем ширину таблицы чтобы влезла в PDF: обрезаем длинные строки
    max_cols = max(len(r) for r in rows)
    # Выравниваем строки по количеству колонок
    for r in rows:
        while len(r) < max_cols:
            r.append("")

    # Обрезаем очень длинные ячейки
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            if len(v) > 120:
                rows[i][j] = v[:117] + "..."

    # Создаём PDF — сначала готовим стили и таблицу, затем подбираем высоту страницы
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    # Базовый формат — A4 или альбом, если много колонок
    base_pagesize = landscape(A4) if max_cols > 5 else A4

    # Стиль таблицы
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
    )

    # reportlab Table требует Paragraph для переноса строк в ячейках
    style_sheet = getSampleStyleSheet()
    cell_style = style_sheet["Normal"]
    cell_style.fontSize = 7
    cell_style.leading = 8
    cell_style.alignment = TA_LEFT

    def _cell(v: str):
        # Экранируем спецсимволы для Paragraph
        esc = str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(esc or " ", cell_style)

    table_data = [[_cell(v) for v in r] for r in rows]

    # Ширина колонок — равномерно (учитываем поля 10мм слева/справа)
    avail_w = base_pagesize[0] - 20 * mm
    col_w = avail_w / max_cols if max_cols else avail_w
    col_widths = [col_w] * max_cols

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(style)

    # Оцениваем высоту таблицы, чтобы убрать лишнее пустое место внизу
    # wrap с большой высотой даёт реальную высоту контента
    try:
        _tw, th = table.wrap(avail_w, 9999 * mm)
    except Exception:
        th = len(rows) * 14  # fallback ~14pt на строку
    title_h = 22  # заголовок + отступ
    needed_h = th + title_h + 20 * mm  # верх+низ поля
    # Минимальная и максимальная высоты
    min_h = 90 * mm
    max_h = base_pagesize[1]
    if needed_h < min_h:
        needed_h = min_h
    # Если таблица маленькая (<=15 строк) и заметно меньше A4 — делаем страницу короче,
    # чтобы не было 2-3x пустого места внизу. Для больших таблиц оставляем A4, чтобы не плодить мелкие страницы.
    if len(rows) <= 15 and needed_h < max_h - 20 * mm:
        pagesize = (base_pagesize[0], needed_h)
    else:
        pagesize = base_pagesize

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=pagesize,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title=xlsx_path.stem,
    )

    story = []
    title_style = style_sheet["Title"]
    title_style.fontSize = 11
    title_style.leading = 14
    story.append(Paragraph(xlsx_path.name, title_style))
    story.append(Spacer(1, 4))
    story.append(table)
    doc.build(story)


# ==============================================================================
#                       ФОНОВЫЕ ПОТОКИ ОБРАБОТКИ
# ==============================================================================


def process_search(config, log_callback, progress_callback, finish_callback):
    """Фоновая функция обработки данных (Массовое сопоставление из файла)."""
    try:
        log_callback(f"Чтение файла: {config['input_file']}")
        wb_list = openpyxl.load_workbook(config["input_file"])
        ws_list = wb_list.active

        start_row = 2 if config["has_header"] else 1
        id_to_rows = defaultdict(list)

        for row_idx in range(start_row, ws_list.max_row + 1):
            cell_val = ws_list.cell(
                row=row_idx, column=config["input_id_col"]
            ).value
            cleaned_id = clean_val(cell_val)
            if cleaned_id:
                id_to_rows[cleaned_id].append(row_idx)

        target_ids = set(id_to_rows.keys())
        log_callback(
            f"Найдено уникальных значений для поиска: {len(target_ids)}"
        )

        if not target_ids:
            log_callback("Ошибка: Список искомых значений пуст!")
            finish_callback(
                False, "В исходном файле не найдено данных для поиска."
            )
            return

        acts_path = Path(config["acts_dir"])
        all_files = [
            f for f in acts_path.rglob("*.xlsx") if not f.name.startswith("~$")
        ]
        total_files = len(all_files)
        log_callback(f"Найдено файлов для проверки: {total_files}")

        if total_files == 0:
            log_callback("Ошибка: В указанной папке нет файлов .xlsx")
            finish_callback(False, "В папке поиска нет файлов .xlsx")
            return

        id_to_results = defaultdict(list)
        search_col = config["target_search_col"]
        fetch_col = config["fetch_cell_col"]
        fixed_row = config["fixed_row_num"]
        fixed_col = config["fixed_col_num"]
        result_type = config["result_type"]

        for idx, fpath in enumerate(all_files, 1):
            file_result_val = get_file_based_result(
                fpath, acts_path, result_type
            )
            fixed_val_cache = None

            try:
                wb = openpyxl.load_workbook(
                    fpath, read_only=True, data_only=True
                )
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        if search_col is None:
                            row_cells_cleaned = [
                                clean_val(c) for c in row if c is not None
                            ]
                            row_text = " ".join(
                                c for c in row_cells_cleaned if c
                            )
                        else:
                            if (
                                len(row) >= search_col
                                and row[search_col - 1] is not None
                            ):
                                row_text = clean_val(row[search_col - 1])
                            else:
                                row_text = ""

                        if not row_text:
                            continue

                        for tid in target_ids:
                            if tid in row_text:
                                if result_type == "row_cell_value":
                                    val_to_add = ""
                                    if (
                                        len(row) >= fetch_col
                                        and row[fetch_col - 1] is not None
                                    ):
                                        val_to_add = clean_val(
                                            row[fetch_col - 1]
                                        )
                                    if val_to_add:
                                        id_to_results[tid].append(val_to_add)

                                elif result_type == "fixed_row_value":
                                    if fixed_val_cache is None:
                                        fixed_val_cache = get_fixed_cell_value(
                                            fpath, fixed_row, fixed_col
                                        )
                                    if fixed_val_cache:
                                        id_to_results[tid].append(
                                            fixed_val_cache
                                        )
                                else:
                                    id_to_results[tid].append(file_result_val)
                wb.close()
            except Exception as e:
                log_callback(f"Ошибка чтения {fpath.name}: {e}")

            progress_callback(int((idx / total_files) * 100))

        # Запись результатов
        matched_count = 0
        for tid, row_indexes in id_to_rows.items():
            matched_items = id_to_results.get(tid, [])
            if matched_items:
                matched_count += 1
                unique_items = list(dict.fromkeys(matched_items))
                val_to_write = ", ".join(unique_items)
            else:
                val_to_write = ""

            for r_idx in row_indexes:
                ws_list.cell(
                    row=r_idx, column=config["output_col"], value=val_to_write
                )

        wb_list.save(config["output_file"])
        log_callback(
            f"Успешно! Найдено совпадений: {matched_count} из {len(target_ids)}"
        )
        finish_callback(
            True,
            f"Готово!\nСовпадений: {matched_count} из {len(target_ids)}\nСохранено в: {config['output_file']}",
        )

    except Exception as e:
        log_callback(f"Критическая ошибка: {e}")
        finish_callback(False, str(e))


def process_multi_query_search(
    folder_path: str,
    query_list: list[str],
    progress_callback,
    finish_callback,
    add_result_callback,
):
    """Фоновый поиск списка строк по папке с файлами."""
    try:
        base_dir = Path(folder_path)
        all_files = [
            f for f in base_dir.rglob("*.xlsx") if not f.name.startswith("~$")
        ]
        total_files = len(all_files)

        if total_files == 0:
            finish_callback(False, "В выбранной папке нет файлов .xlsx")
            return

        queries = [
            (q_orig, q_orig.lower()) for q_orig in query_list if q_orig.strip()
        ]
        found_files_set = set()

        for idx, fpath in enumerate(all_files, 1):
            matched_queries = set()
            try:
                wb = openpyxl.load_workbook(
                    fpath, read_only=True, data_only=True
                )
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_cells_cleaned = [
                            clean_val(c).lower() for c in row if c is not None
                        ]
                        row_text = " ".join(c for c in row_cells_cleaned if c)

                        for q_orig, q_low in queries:
                            if q_low in row_text:
                                matched_queries.add(q_orig)

                wb.close()
            except Exception:
                pass

            if matched_queries:
                found_files_set.add(str(fpath.resolve()))
                try:
                    rel_path = str(fpath.relative_to(base_dir).parent)
                    if rel_path == ".":
                        rel_path = "Корень папки"
                except ValueError:
                    rel_path = ""

                add_result_callback(
                    (
                        fpath.name,
                        ", ".join(sorted(matched_queries)),
                        rel_path,
                        str(fpath.resolve()),
                    )
                )

            progress_callback(int((idx / total_files) * 100))

        finish_callback(
            True,
            f"Поиск завершен!\nНайдено файлов с совпадениями: {len(found_files_set)} (из {total_files} проверенных).",
        )

    except Exception as e:
        finish_callback(False, f"Ошибка поиска: {e}")


def process_hash_calculation(
    folder_path: str,
    output_excel: str,
    algorithm: str,
    log_callback,
    progress_callback,
    finish_callback,
):
    """Фоновое вычисление хешей файлов и запись только 2 колонок (Имя и Хеш)."""
    try:
        base_dir = Path(folder_path)
        log_callback(f"Сканирование папки: {base_dir}")

        all_files = [
            p
            for p in base_dir.rglob("*")
            if p.is_file() and not p.name.startswith("~$")
        ]
        total_files = len(all_files)

        if total_files == 0:
            finish_callback(False, "В выбранной папке не найдено файлов.")
            return

        log_callback(f"Найдено файлов для хеширования: {total_files}")
        log_callback(f"Выбран алгоритм: {algorithm.upper()}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Хеши файлов"

        headers = ["Имя файла", f"Хеш-сумма ({algorithm.upper()})"]
        ws.append(headers)

        processed_count = 0
        error_count = 0

        for idx, fpath in enumerate(all_files, 1):
            try:
                file_hash = calculate_file_hash(fpath, algorithm)
                ws.append([fpath.name, file_hash])
                processed_count += 1
            except Exception as e:
                error_count += 1
                log_callback(f"Ошибка при обработке {fpath.name}: {e}")

            if idx % 10 == 0 or idx == total_files:
                progress_callback(int((idx / total_files) * 100))

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 68

        wb.save(output_excel)
        log_callback(
            f"Успешно сохранено: {processed_count} записей в {output_excel}"
        )

        msg = f"Хеширование успешно завершено!\n\nОбработано файлов: {processed_count}"
        if error_count > 0:
            msg += f"\nОшибок доступа: {error_count}"
        msg += f"\n\nРезультат сохранен в:\n{output_excel}"

        finish_callback(True, msg)

    except Exception as e:
        log_callback(f"Критическая ошибка: {e}")
        finish_callback(False, str(e))


def extract_column_values(filepath: str, col_idx: int, has_header: bool) -> list[str]:
    """Извлекает список очищенных значений из заданной колонки активного листа."""
    vals = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    start_row = 2 if has_header else 1
    for row in ws.iter_rows(
        min_row=start_row, min_col=col_idx, max_col=col_idx, values_only=True
    ):
        if row and row[0] is not None:
            c = clean_val(row[0])
            if c:
                vals.append(c)
    wb.close()
    return vals


def process_column_diff(config, finish_callback):
    """Фоновое вычисление разницы колонок двух файлов с формированием Side-by-Side diff."""
    try:
        f1 = config["file1"]
        f2 = config["file2"]
        col1 = config["col1"]
        col2 = config["col2"]
        hdr1 = config["has_header1"]
        hdr2 = config["has_header2"]

        vals1 = extract_column_values(f1, col1, hdr1)
        vals2 = extract_column_values(f2, col2, hdr2)

        set1 = set(vals1)
        set2 = set(vals2)

        only_in_1 = [x for x in dict.fromkeys(vals1) if x not in set2]
        only_in_2 = [x for x in dict.fromkeys(vals2) if x not in set1]
        common = [x for x in dict.fromkeys(vals1) if x in set2]

        # Формируем объединенный сравнительный список (Side-by-Side Diff)
        # 1. Сначала общие элементы
        # 2. Затем элементы только из файла 1
        # 3. Затем элементы только из файла 2
        side_by_side_rows = []
        for v in common:
            side_by_side_rows.append((v, v, "Общее", "equal"))
        for v in only_in_1:
            side_by_side_rows.append((v, "—", "Только в Файле 1", "only1"))
        for v in only_in_2:
            side_by_side_rows.append(("—", v, "Только в Файле 2", "only2"))

        results = {
            "total_f1_rows": len(vals1),
            "total_f2_rows": len(vals2),
            "unique_f1": len(set1),
            "unique_f2": len(set2),
            "only_in_1": only_in_1,
            "only_in_2": only_in_2,
            "common": common,
            "side_by_side": side_by_side_rows,
        }
        finish_callback(True, results)
    except Exception as e:
        finish_callback(False, str(e))


def process_xlsx_to_pdf(
    source_dir: str,
    output_dir: str,
    naming_mode: str,
    log_callback,
    progress_callback,
    finish_callback,
    keep_structure: bool = False,
    engine: str = "auto",
    google_credentials: str = "credentials.json",
    google_token: str = "token.json",
):
    """Фоновая конвертация папки с xlsx в pdf с выбором схемы именования.

    naming_mode:
        'filename'      — имя pdf = имя исходного файла (stem)
        'parent_folder' — имя pdf = имя родительской папки
    keep_structure:
        True — сохранять структуру подпапок (out/подпапка/файл.pdf)
        False — складывать все PDF плоско в одну папку
    engine:
        'auto' — LibreOffice -> Excel -> reportlab (приоритет)
        'libre' — только LibreOffice
        'excel' — только Microsoft Excel (COM)
        'reportlab' — только reportlab (упрощённо)
        'google' — Google Sheets (Drive API, требует credentials.json)
    """
    try:
        src = Path(source_dir)
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)

        all_files = [p for p in src.rglob("*.xlsx") if not p.name.startswith("~$") and p.is_file()]
        # Также поддержим .xls (если найдутся)
        all_files += [p for p in src.rglob("*.xls") if not p.name.startswith("~$") and p.is_file() and p.suffix.lower() == ".xls"]
        # Дедупликация (на случай пересечения масок)
        all_files = sorted(set(all_files))

        total = len(all_files)
        if total == 0:
            finish_callback(False, "В выбранной папке нет файлов .xlsx / .xls")
            return

        log_callback(f"Найдено файлов для конвертации: {total}")
        log_callback(f"Исходная папка: {src}")
        log_callback(f"Выходная папка: {out}")
        log_callback(f"Схема именования: {'Имя файла' if naming_mode == 'filename' else 'Имя родительской папки'}")
        log_callback(f"Структура папок: {'сохранять' if keep_structure else 'плоско (в одну папку)'}")

        soffice = find_soffice_executable()
        use_libre = False
        use_excel = False
        use_google = False
        excel_app = None
        drive_service = None
        google_creds = None
        engine = (engine or "auto").lower()
        log_callback(f"Выбран движок: {engine}")

        if engine == "libre":
            if soffice:
                use_libre = True
                log_callback(f"Используется LibreOffice: {soffice} — точная конвертация")
            else:
                finish_callback(False, "LibreOffice не найден.\nУстановите LibreOffice с https://www.libreoffice.org/\nили выберите другой движок (Excel/reportlab).")
                return
        elif engine == "excel":
            if not _is_excel_available():
                finish_callback(False, "Microsoft Excel не найден или недоступен через COM.\nУбедитесь, что Excel установлен на Windows.")
                return
            try:
                import win32com.client
                import pythoncom

                pythoncom.CoInitialize()
                excel_app = win32com.client.DispatchEx("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                try:
                    excel_app.ScreenUpdating = False
                except Exception:
                    pass
                use_excel = True
                log_callback("Используется Microsoft Excel (COM) — точная конвертация")
            except Exception as e:
                log_callback(f"Не удалось запустить Excel: {e}")
                try:
                    if excel_app is not None:
                        excel_app.Quit()
                except Exception:
                    pass
                excel_app = None
                finish_callback(False, f"Ошибка запуска Excel: {e}")
                return
        elif engine == "reportlab":
            try:
                import reportlab  # noqa: F401

                log_callback("Используется reportlab (упрощённая табличная конвертация)")
            except ImportError:
                finish_callback(False, "Не установлен reportlab.\nВыполните: pip install reportlab")
                return
        elif engine == "google":
            try:
                drive_service, google_creds = _get_google_drive_service(google_credentials, google_token)
                use_google = True
                log_callback("Используется Google Sheets (Drive API) — конвертация через Google (A4, альбом, fitw)")
            except Exception as e:
                finish_callback(False, f"Ошибка Google Drive: {e}\nПроверьте credentials.json и доступ к интернету.")
                return
        else:  # auto
            if soffice:
                use_libre = True
                log_callback(f"Используется LibreOffice: {soffice} — точная конвертация (приоритет)")
            elif _is_excel_available():
                try:
                    import win32com.client
                    import pythoncom

                    pythoncom.CoInitialize()
                    excel_app = win32com.client.DispatchEx("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    try:
                        excel_app.ScreenUpdating = False
                    except Exception:
                        pass
                    use_excel = True
                    log_callback("LibreOffice не найден — будет использован Microsoft Excel (COM) — точная конвертация")
                except Exception as e:
                    log_callback(f"Microsoft Excel найден, но не удалось запустить: {e}")
                    try:
                        if excel_app is not None:
                            excel_app.Quit()
                    except Exception:
                        pass
                    excel_app = None
                    use_excel = False
                    try:
                        import reportlab  # noqa: F401

                        log_callback("Будет использован reportlab (упрощённая табличная конвертация)")
                    except ImportError:
                        finish_callback(
                            False,
                            "Не найден LibreOffice (soffice), недоступен Microsoft Excel и не установлен reportlab.\n"
                            "Установите LibreOffice с https://www.libreoffice.org/\n"
                            "или убедитесь что установлен Microsoft Excel,\n"
                            "или выполните: pip install reportlab",
                        )
                        return
            else:
                try:
                    import reportlab  # noqa: F401

                    log_callback("LibreOffice и Excel не найдены — будет использован reportlab (упрощённая табличная конвертация)")
                except ImportError:
                    finish_callback(
                        False,
                        "Не найден LibreOffice (soffice), недоступен Microsoft Excel и не установлен reportlab.\n"
                        "Установите LibreOffice с https://www.libreoffice.org/\n"
                        "или убедитесь что установлен Microsoft Excel,\n"
                        "или выполните: pip install reportlab",
                    )
                    return

        # Для плоского режима — один общий set, для структуры — по папкам
        used_names_per_dir: dict[str, set[str]] = {}
        success = 0
        errors = 0

        for idx, fpath in enumerate(all_files, 1):
            try:
                # --- целевая директория ---
                if keep_structure:
                    try:
                        rel_parent = fpath.parent.relative_to(src)
                    except ValueError:
                        rel_parent = Path(".")
                    if str(rel_parent) == ".":
                        target_dir = out
                    else:
                        target_dir = out / rel_parent
                    target_dir.mkdir(parents=True, exist_ok=True)
                else:
                    target_dir = out

                # Определяем базовое имя
                if naming_mode == "parent_folder":
                    base = fpath.parent.name.strip()
                    # Если файл лежит прямо в корне исходной папки — parent == src.name, это ок
                    # Если по какой-то причине имя пустое — fallback на stem
                    if not base or base in (".", ""):
                        base = fpath.stem
                else:
                    base = fpath.stem

                base = sanitize_filename(base)
                # Уникализация имени в пределах целевой папки
                dir_key = str(target_dir).lower()
                if dir_key not in used_names_per_dir:
                    used_names_per_dir[dir_key] = set()
                used = used_names_per_dir[dir_key]
                candidate = base
                counter = 1
                # Учитываем уже занятые имена в этой сессии и существующие файлы на диске в target_dir
                while candidate.lower() in used or (target_dir / f"{candidate}.pdf").exists():
                    candidate = f"{base}_{counter}"
                    counter += 1
                used.add(candidate.lower())
                target_pdf = target_dir / f"{candidate}.pdf"

                if use_libre:
                    # Конвертация через LibreOffice (промежуточный файл в tmp)
                    interim = _convert_single_xlsx_to_pdf_libreoffice(soffice, fpath, out)
                    try:
                        if target_pdf.exists():
                            target_pdf.unlink()
                        # interim находится во временной папке — перемещаем
                        shutil.move(str(interim), str(target_pdf))
                        # Удаляем временную папку
                        try:
                            tmp_root = interim.parent
                            if tmp_root.name.startswith("xlsx2pdf_"):
                                shutil.rmtree(tmp_root, ignore_errors=True)
                        except Exception:
                            pass
                    except Exception:
                        # fallback: пробуем rename
                        try:
                            interim.rename(target_pdf)
                        except Exception:
                            shutil.copy2(str(interim), str(target_pdf))
                    # Красивое имя для лога: при структуре показываем подпапку
                    try:
                        rel_out = str(target_pdf.relative_to(out))
                    except ValueError:
                        rel_out = target_pdf.name
                    log_callback(f"[{idx}/{total}] OK: {fpath.name} -> {rel_out}")
                elif use_excel:
                    _convert_single_xlsx_to_pdf_excel(excel_app, fpath, target_pdf)
                    try:
                        rel_out = str(target_pdf.relative_to(out))
                    except ValueError:
                        rel_out = target_pdf.name
                    log_callback(f"[{idx}/{total}] OK (Excel): {fpath.name} -> {rel_out}")
                elif use_google:
                    _convert_single_xlsx_to_pdf_google(drive_service, google_creds, fpath, target_pdf)
                    try:
                        rel_out = str(target_pdf.relative_to(out))
                    except ValueError:
                        rel_out = target_pdf.name
                    log_callback(f"[{idx}/{total}] OK (Google): {fpath.name} -> {rel_out}")
                else:
                    _convert_single_xlsx_to_pdf_reportlab(fpath, target_pdf)
                    try:
                        rel_out2 = str(target_pdf.relative_to(out))
                    except ValueError:
                        rel_out2 = target_pdf.name
                    log_callback(f"[{idx}/{total}] OK (reportlab): {fpath.name} -> {rel_out2}")

                success += 1
            except Exception as e:
                errors += 1
                log_callback(f"[{idx}/{total}] Ошибка {fpath.name}: {e}")

            progress_callback(int((idx / total) * 100))

        # Корректно закрываем Excel, если использовали
        if excel_app is not None:
            try:
                excel_app.Quit()
            except Exception:
                pass
            try:
                import pythoncom

                pythoncom.CoUninitialize()
            except Exception:
                pass

        msg = f"Конвертация завершена!\nУспешно: {success} из {total}"
        if errors:
            msg += f"\nОшибок: {errors}"
        msg += f"\nПапка с PDF: {out}"
        if use_libre:
            msg += "\nДвижок: LibreOffice"
        elif use_excel:
            msg += "\nДвижок: Microsoft Excel"
        elif use_google:
            msg += "\nДвижок: Google Sheets"
        else:
            msg += "\nДвижок: reportlab (упрощённо)"
        finish_callback(True, msg)

    except Exception as e:
        # Пытаемся закрыть Excel даже при критической ошибке
        try:
            if "excel_app" in locals() and excel_app is not None:
                excel_app.Quit()
        except Exception:
            pass
        try:
            import pythoncom

            pythoncom.CoUninitialize()
        except Exception:
            pass
        log_callback(f"Критическая ошибка: {e}")
        finish_callback(False, str(e))


# ==============================================================================
#                                GUI ИНТЕРФЕЙС
# ==============================================================================


class ExcelFinderApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Excel Tool - Match, Search, Hash & Diff")
        self.geometry("860x920")
        self.minsize(760, 820)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_bulk = ttk.Frame(notebook)
        self.tab_single = ttk.Frame(notebook)
        self.tab_hash = ttk.Frame(notebook)
        self.tab_diff = ttk.Frame(notebook)
        self.tab_pdf = ttk.Frame(notebook)

        notebook.add(self.tab_bulk, text="  Массовое сопоставление  ")
        notebook.add(self.tab_single, text="  Поиск по списку строк  ")
        notebook.add(self.tab_hash, text="  Хеш-генератор  ")
        notebook.add(self.tab_diff, text="  Сравнение колонок (Diff)  ")
        notebook.add(self.tab_pdf, text="  Конвертер XLSX -> PDF  ")

        self._build_tab_bulk()
        self._build_tab_single()
        self._build_tab_hash()
        self._build_tab_diff()
        self._build_tab_pdf()

    def _add_tab_description(self, parent, text: str):
        """Добавляет голубой блок-описание вверху вкладки простым языком."""
        wrapper = tk.Frame(parent, bg="#EAF0FF", highlightbackground="#C5D3EA", highlightthickness=1)
        wrapper.pack(fill="x", padx=10, pady=(10, 4))
        # иконка ℹ
        inner = tk.Frame(wrapper, bg="#EAF0FF")
        inner.pack(fill="x", padx=10, pady=8)
        tk.Label(inner, text="ℹ", bg="#EAF0FF", fg="#2F5AA0", font=("Segoe UI", 11, "bold")).pack(side="left", padx=(0, 8), anchor="nw")
        tk.Label(
            inner,
            text=text,
            bg="#EAF0FF",
            fg="#243656",
            font=("Segoe UI", 9),
            wraplength=760,
            justify="left",
            anchor="w",
        ).pack(side="left", fill="both", expand=True, anchor="w")

    # --------------------------------------------------------------------------
    #                     ВКЛАДКА 1: МАССОВОЕ СОПОСТАВЛЕНИЕ
    # --------------------------------------------------------------------------
    def _build_tab_bulk(self):
        pad_opts = {"padx": 10, "pady": 4}

        self._add_tab_description(
            self.tab_bulk,
            "Вы можете взять список ID/номеров из вашего исходного Excel и автоматически найти каждое значение во всех Excel-файлах внутри выбранной папки (включая подпапки). "
            "Найденное записывается в указанную колонку нового файла. Удобно, когда нужно сопоставить сотни номеров с актами или накладными без ручного поиска.",
        )

        frame_files = ttk.LabelFrame(
            self.tab_bulk, text=" 1. Файлы и папки ", padding=10
        )
        frame_files.pack(fill="x", **pad_opts)

        ttk.Label(frame_files, text="Исходный Excel файл:").grid(
            row=0, column=0, sticky="w"
        )
        self.ent_input_file = ttk.Entry(frame_files, width=42)
        self.ent_input_file.grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(frame_files, text="Обзор...", command=self._browse_input).grid(
            row=0, column=2
        )

        ttk.Label(frame_files, text="Папка для поиска:").grid(
            row=1, column=0, sticky="w"
        )
        self.ent_acts_dir = ttk.Entry(frame_files, width=42)
        self.ent_acts_dir.grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(
            frame_files, text="Обзор...", command=self._browse_acts_dir
        ).grid(row=1, column=2)

        ttk.Label(frame_files, text="Куда сохранить результат:").grid(
            row=2, column=0, sticky="w"
        )
        self.ent_output_file = ttk.Entry(frame_files, width=42)
        self.ent_output_file.grid(row=2, column=1, padx=5, pady=2)
        ttk.Button(
            frame_files, text="Обзор...", command=self._browse_output
        ).grid(row=2, column=2)

        frame_settings = ttk.LabelFrame(
            self.tab_bulk,
            text=" 2. Настройки исходного файла и поиска ",
            padding=10,
        )
        frame_settings.pack(fill="x", **pad_opts)

        ttk.Label(
            frame_settings, text="Номер колонки с искомыми ID (напр. 1=A):"
        ).grid(row=0, column=0, sticky="w", pady=2)
        self.spn_input_col = ttk.Spinbox(
            frame_settings, from_=1, to=100, width=8
        )
        self.spn_input_col.set(1)
        self.spn_input_col.grid(row=0, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(
            frame_settings, text="Номер колонки для записи результата:"
        ).grid(row=1, column=0, sticky="w", pady=2)
        self.spn_output_col = ttk.Spinbox(
            frame_settings, from_=1, to=100, width=8
        )
        self.spn_output_col.set(2)
        self.spn_output_col.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(
            frame_settings,
            text="Искать данные в конкретной колонке (пусто — искать по всем ячейкам):",
        ).grid(row=2, column=0, sticky="w", pady=2)
        self.ent_search_col = ttk.Entry(frame_settings, width=9)
        self.ent_search_col.grid(row=2, column=1, sticky="w", padx=5, pady=2)

        self.var_has_header = tk.BooleanVar(value=False)
        self.chk_header = ttk.Checkbutton(
            frame_settings,
            text="В исходном файле есть заголовки (пропустить 1-ю строку)",
            variable=self.var_has_header,
        )
        self.chk_header.grid(
            row=3, column=0, columnspan=2, sticky="w", pady=(4, 0)
        )

        frame_res = ttk.LabelFrame(
            self.tab_bulk, text=" 3. Формат результата ", padding=10
        )
        frame_res.pack(fill="x", **pad_opts)

        self.var_res_type = tk.StringVar(value="last_number")

        res_options = [
            (
                "last_number",
                "Последний номер из названия файла (например, 262075388 из 'act-income-mp-262075388.xlsx')",
            ),
            ("full_filename", "Полное имя файла (act-income-mp-262075388.xlsx)"),
            ("filename_stem", "Имя файла без расширения (act-income-mp-262075388)"),
            ("relative_path", "Путь к файлу с подпапками (D:/acts/act-income-mp-262075388.xlsx)"),
            (
                "row_cell_value",
                "Значение из совпавшей строки (укажите колонку):",
            ),
            (
                "fixed_row_value",
                "Значение из фиксированной строки (укажите строку и колонку):",
            ),
        ]

        for idx, (val, text) in enumerate(res_options):
            rb = ttk.Radiobutton(
                frame_res,
                text=text,
                value=val,
                variable=self.var_res_type,
                command=self._update_input_states,
            )
            rb.grid(row=idx * 2, column=0, sticky="w", pady=1)

            if val == "row_cell_value":
                self.frame_match_row = ttk.Frame(frame_res)
                self.frame_match_row.grid(
                    row=idx * 2 + 1, column=0, sticky="w", padx=25, pady=2
                )
                ttk.Label(
                    self.frame_match_row, text="Номер колонки в строке:"
                ).pack(side="left")
                self.spn_fetch_col = ttk.Spinbox(
                    self.frame_match_row, from_=1, to=100, width=6
                )
                self.spn_fetch_col.set(3)
                self.spn_fetch_col.pack(side="left", padx=5)

            if val == "fixed_row_value":
                self.frame_fixed_row = ttk.Frame(frame_res)
                self.frame_fixed_row.grid(
                    row=idx * 2 + 1, column=0, sticky="w", padx=25, pady=2
                )
                ttk.Label(self.frame_fixed_row, text="Номер строки:").pack(
                    side="left"
                )
                self.spn_fixed_row = ttk.Spinbox(
                    self.frame_fixed_row, from_=1, to=1000, width=6
                )
                self.spn_fixed_row.set(1)
                self.spn_fixed_row.pack(side="left", padx=5)

                ttk.Label(self.frame_fixed_row, text="Номер колонки:").pack(
                    side="left", padx=(10, 0)
                )
                self.spn_fixed_col = ttk.Spinbox(
                    self.frame_fixed_row, from_=1, to=100, width=6
                )
                self.spn_fixed_col.set(1)
                self.spn_fixed_col.pack(side="left", padx=5)

        self._update_input_states()

        self.btn_run = ttk.Button(
            self.tab_bulk, text="▶ Запустить поиск", command=self._start_search
        )
        self.btn_run.pack(fill="x", padx=10, pady=6)

        self.progress = ttk.Progressbar(self.tab_bulk, mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=2)

        self.txt_log = tk.Text(
            self.tab_bulk, height=6, wrap="word", font=("Consolas", 9)
        )
        self.txt_log.pack(fill="both", expand=True, padx=10, pady=(4, 8))

    def _update_input_states(self):
        res_type = self.var_res_type.get()
        self.spn_fetch_col.config(
            state="normal" if res_type == "row_cell_value" else "disabled"
        )
        state_fixed = (
            "normal" if res_type == "fixed_row_value" else "disabled"
        )
        self.spn_fixed_row.config(state=state_fixed)
        self.spn_fixed_col.config(state=state_fixed)

    def _browse_input(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if path:
            self.ent_input_file.delete(0, tk.END)
            self.ent_input_file.insert(0, path)
            if not self.ent_output_file.get():
                p = Path(path)
                self.ent_output_file.insert(
                    0, str(p.parent / f"{p.stem}_result.xlsx")
                )

    def _browse_acts_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.ent_acts_dir.delete(0, tk.END)
            self.ent_acts_dir.insert(0, path)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.ent_output_file.delete(0, tk.END)
            self.ent_output_file.insert(0, path)

    def _start_search(self):
        in_file = self.ent_input_file.get().strip()
        acts_dir = self.ent_acts_dir.get().strip()
        out_file = self.ent_output_file.get().strip()

        if not in_file or not os.path.isfile(in_file):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректный исходный файл Excel!"
            )
            return
        if not acts_dir or not os.path.isdir(acts_dir):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректную папку!"
            )
            return
        if not out_file:
            messagebox.showwarning(
                "Предупреждение", "Укажите путь для сохранения результата!"
            )
            return

        search_col_val = self.ent_search_col.get().strip()
        target_search_col = (
            int(search_col_val) if search_col_val.isdigit() else None
        )

        config = {
            "input_file": in_file,
            "acts_dir": acts_dir,
            "output_file": out_file,
            "input_id_col": int(self.spn_input_col.get()),
            "output_col": int(self.spn_output_col.get()),
            "has_header": self.var_has_header.get(),
            "target_search_col": target_search_col,
            "result_type": self.var_res_type.get(),
            "fetch_cell_col": int(self.spn_fetch_col.get()),
            "fixed_row_num": int(self.spn_fixed_row.get()),
            "fixed_col_num": int(self.spn_fixed_col.get()),
        }

        self.txt_log.delete(1.0, tk.END)
        self.progress["value"] = 0
        self.btn_run.config(state="disabled")

        threading.Thread(
            target=process_search,
            args=(
                config,
                lambda t: (
                    self.txt_log.insert(tk.END, t + "\n"),
                    self.txt_log.see(tk.END),
                ),
                lambda v: self.progress.config(value=v),
                lambda s, m: (
                    self.btn_run.config(state="normal"),
                    (
                        messagebox.showinfo("Успех", m)
                        if s
                        else messagebox.showerror("Ошибка", m)
                    ),
                ),
            ),
            daemon=True,
        ).start()

    # --------------------------------------------------------------------------
    #                     ВКЛАДКА 2: ПОИСК ПО СПИСКУ СТРОК
    # --------------------------------------------------------------------------
    def _build_tab_single(self):
        pad_opts = {"padx": 10, "pady": 5}

        self._add_tab_description(
            self.tab_single,
            "Вставьте список значений (через запятую или каждый с новой строки), укажите папку с Excel файлами — программа найдёт все файлы, где встречается хотя бы одно из значений. "
            "Результат — таблица с именем, подпапкой и полным путём. Двойной клик открывает файл, можно скопировать найденные файлы в одну папку или выгрузить пути в TXT/буфер.",
        )

        frame_input = ttk.LabelFrame(
            self.tab_single,
            text=" 1. Параметры поиска и список строк ",
            padding=10,
        )
        frame_input.pack(fill="x", **pad_opts)

        f_folder = ttk.Frame(frame_input)
        f_folder.pack(fill="x", pady=(0, 5))
        ttk.Label(f_folder, text="Папка с Excel файлами:").pack(
            side="left", padx=(0, 5)
        )
        self.ent_single_dir = ttk.Entry(f_folder)
        self.ent_single_dir.pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(
            f_folder, text="Обзор...", command=self._browse_single_dir
        ).pack(side="right")

        ttk.Label(
            frame_input,
            text="Искомые строки (каждый с новой строки или через запятую):",
        ).pack(anchor="w", pady=(5, 2))

        self.txt_queries = tk.Text(
            frame_input, height=5, font=("Consolas", 10), wrap="word"
        )
        self.txt_queries.pack(fill="x", pady=2)

        self._setup_text_widget_clipboard(self.txt_queries)

        self.btn_single_run = ttk.Button(
            frame_input,
            text="🔍 Найти файлы по списку",
            command=self._start_single_search,
        )
        self.btn_single_run.pack(fill="x", pady=(5, 0))

        self.single_progress = ttk.Progressbar(
            self.tab_single, mode="determinate"
        )
        self.single_progress.pack(fill="x", padx=10, pady=(4, 6))

        frame_results = ttk.LabelFrame(
            self.tab_single, text=" 2. Результаты поиска ", padding=8
        )
        frame_results.pack(fill="both", expand=True, **pad_opts)

        columns = ("filename", "query", "folder", "fullpath")
        self.tree_results = ttk.Treeview(
            frame_results, columns=columns, show="headings", selectmode="browse"
        )

        self.tree_results.heading("filename", text="Имя файла")
        self.tree_results.heading("query", text="Найденный текст")
        self.tree_results.heading("folder", text="Подпапка")
        self.tree_results.heading("fullpath", text="Полный путь")

        self.tree_results.column("filename", width=190, anchor="w")
        self.tree_results.column("query", width=140, anchor="w")
        self.tree_results.column("folder", width=120, anchor="w")
        self.tree_results.column("fullpath", width=220, anchor="w")

        scrollbar = ttk.Scrollbar(
            frame_results, orient="vertical", command=self.tree_results.yview
        )
        self.tree_results.configure(yscrollcommand=scrollbar.set)

        self.tree_results.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree_results.bind("<Double-1>", self._open_selected_file)

        frame_actions = ttk.Frame(self.tab_single)
        frame_actions.pack(fill="x", padx=10, pady=(0, 10))

        self.lbl_single_status = ttk.Label(
            frame_actions, text="Готов к поиску (двойной клик откроет файл)"
        )
        self.lbl_single_status.pack(side="left")

        ttk.Button(
            frame_actions,
            text="📂 Скопировать файлы в папку...",
            command=self._copy_files_to_folder,
        ).pack(side="right", padx=3)
        ttk.Button(
            frame_actions,
            text="📋 Скопировать пути",
            command=self._copy_results_to_clipboard,
        ).pack(side="right", padx=3)
        ttk.Button(
            frame_actions,
            text="💾 Экспорт в TXT",
            command=self._export_results_to_txt,
        ).pack(side="right", padx=3)

    # --------------------------------------------------------------------------
    #                     ВКЛАДКА 3: ХЕШ-СУММЫ ФАЙЛОВ
    # --------------------------------------------------------------------------
    def _build_tab_hash(self):
        pad_opts = {"padx": 10, "pady": 5}

        self._add_tab_description(
            self.tab_hash,
            "Вы можете считать контрольные суммы (хеши) всех файлов в папке и сохраняет таблицу «имя файла — хеш» в Excel. "
            "Помогает проверить, не изменился ли файл, и найти дубликаты. Выберите алгоритм: SHA-256 (самый надёжный), MD5 или SHA-1.",
        )

        frame_params = ttk.LabelFrame(
            self.tab_hash,
            text=" 1. Параметры сканирования и экспорта ",
            padding=10,
        )
        frame_params.pack(fill="x", **pad_opts)

        ttk.Label(frame_params, text="Папка с файлами:").grid(
            row=0, column=0, sticky="w"
        )
        self.ent_hash_dir = ttk.Entry(frame_params, width=42)
        self.ent_hash_dir.grid(row=0, column=1, padx=5, pady=3)
        ttk.Button(
            frame_params, text="Обзор...", command=self._browse_hash_dir
        ).grid(row=0, column=2)

        ttk.Label(frame_params, text="Файл для сохранения (.xlsx):").grid(
            row=1, column=0, sticky="w"
        )
        self.ent_hash_output = ttk.Entry(frame_params, width=42)
        self.ent_hash_output.grid(row=1, column=1, padx=5, pady=3)
        ttk.Button(
            frame_params, text="Обзор...", command=self._browse_hash_output
        ).grid(row=1, column=2)

        frame_algo = ttk.Frame(frame_params)
        frame_algo.grid(
            row=2, column=0, columnspan=3, sticky="w", pady=(8, 2)
        )

        ttk.Label(frame_algo, text="Алгоритм хеширования:").pack(
            side="left", padx=(0, 10)
        )
        self.var_hash_algo = tk.StringVar(value="sha256")

        ttk.Radiobutton(
            frame_algo,
            text="SHA-256 (рекомендуется)",
            value="sha256",
            variable=self.var_hash_algo,
        ).pack(side="left", padx=5)
        ttk.Radiobutton(
            frame_algo, text="MD5", value="md5", variable=self.var_hash_algo
        ).pack(side="left", padx=5)
        ttk.Radiobutton(
            frame_algo, text="SHA-1", value="sha1", variable=self.var_hash_algo
        ).pack(side="left", padx=5)

        self.btn_hash_run = ttk.Button(
            self.tab_hash,
            text="▶ Вычислить хеши и сохранить в Excel",
            command=self._start_hash_calculation,
        )
        self.btn_hash_run.pack(fill="x", padx=10, pady=6)

        self.hash_progress = ttk.Progressbar(self.tab_hash, mode="determinate")
        self.hash_progress.pack(fill="x", padx=10, pady=2)

        frame_log = ttk.LabelFrame(
            self.tab_hash, text=" 2. Журнал обработки ", padding=8
        )
        frame_log.pack(fill="both", expand=True, **pad_opts)

        self.txt_hash_log = tk.Text(
            frame_log, height=10, wrap="word", font=("Consolas", 9)
        )
        self.txt_hash_log.pack(fill="both", expand=True)

    def _browse_hash_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.ent_hash_dir.delete(0, tk.END)
            self.ent_hash_dir.insert(0, path)
            if not self.ent_hash_output.get():
                p = Path(path)
                self.ent_hash_output.insert(
                    0, str(p.parent / f"{p.name}_file_hashes.xlsx")
                )

    def _browse_hash_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.ent_hash_output.delete(0, tk.END)
            self.ent_hash_output.insert(0, path)

    def _start_hash_calculation(self):
        folder = self.ent_hash_dir.get().strip()
        out_excel = self.ent_hash_output.get().strip()
        algo = self.var_hash_algo.get()

        if not folder or not os.path.isdir(folder):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректную папку с файлами!"
            )
            return
        if not out_excel:
            messagebox.showwarning(
                "Предупреждение",
                "Укажите путь для сохранения выходного Excel файла!",
            )
            return

        self.txt_hash_log.delete(1.0, tk.END)
        self.hash_progress["value"] = 0
        self.btn_hash_run.config(state="disabled")

        threading.Thread(
            target=process_hash_calculation,
            args=(
                folder,
                out_excel,
                algo,
                lambda t: (
                    self.txt_hash_log.insert(tk.END, t + "\n"),
                    self.txt_log.see(tk.END),
                ),
                lambda v: self.hash_progress.config(value=v),
                lambda s, m: (
                    self.btn_hash_run.config(state="normal"),
                    (
                        messagebox.showinfo("Успех", m)
                        if s
                        else messagebox.showerror("Ошибка", m)
                    ),
                ),
            ),
            daemon=True,
        ).start()

    # --------------------------------------------------------------------------
    #                     ВКЛАДКА 4: СРАВНЕНИЕ КОЛОНОК (DIFF)
    # --------------------------------------------------------------------------
    def _build_tab_diff(self):
        pad_opts = {"padx": 10, "pady": 4}

        self._add_tab_description(
            self.tab_diff,
            "Вы можете сравнить два списка из указанных колонок двух Excel-файлов. Показывает, что есть только в первом файле, только во втором и что совпадает. "
            "Есть два вида просмотра — наглядная таблица Side-by-Side с подсветкой (зелёный — совпало, красный/оранжевый — различия) и фильтр по категориям. Можно сохранить отчёт в Excel или TXT.",
        )

        # 1. Параметры файлов и колонок
        frame_inputs = ttk.LabelFrame(
            self.tab_diff, text=" 1. Выбор файлов и колонок для сравнения ", padding=10
        )
        frame_inputs.pack(fill="x", **pad_opts)

        # Файл 1
        f1_row = ttk.Frame(frame_inputs)
        f1_row.pack(fill="x", pady=2)
        ttk.Label(f1_row, text="Файл 1:", width=8).pack(side="left")
        self.ent_diff_file1 = ttk.Entry(f1_row, width=38)
        self.ent_diff_file1.pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(f1_row, text="Обзор...", command=self._browse_diff_file1).pack(
            side="left", padx=2
        )
        ttk.Label(f1_row, text="Колонка:").pack(side="left", padx=(8, 2))
        self.spn_diff_col1 = ttk.Spinbox(f1_row, from_=1, to=100, width=5)
        self.spn_diff_col1.set(1)
        self.spn_diff_col1.pack(side="left", padx=2)

        self.var_diff_hdr1 = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            f1_row, text="Заголовок", variable=self.var_diff_hdr1
        ).pack(side="left", padx=6)

        # Файл 2
        f2_row = ttk.Frame(frame_inputs)
        f2_row.pack(fill="x", pady=4)
        ttk.Label(f2_row, text="Файл 2:", width=8).pack(side="left")
        self.ent_diff_file2 = ttk.Entry(f2_row, width=38)
        self.ent_diff_file2.pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(f2_row, text="Обзор...", command=self._browse_diff_file2).pack(
            side="left", padx=2
        )
        ttk.Label(f2_row, text="Колонка:").pack(side="left", padx=(8, 2))
        self.spn_diff_col2 = ttk.Spinbox(f2_row, from_=1, to=100, width=5)
        self.spn_diff_col2.set(1)
        self.spn_diff_col2.pack(side="left", padx=2)

        self.var_diff_hdr2 = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            f2_row, text="Заголовок", variable=self.var_diff_hdr2
        ).pack(side="left", padx=6)

        # Кнопка запуска
        self.btn_diff_run = ttk.Button(
            frame_inputs,
            text="⚖ Сравнить колонки",
            command=self._start_column_diff,
        )
        self.btn_diff_run.pack(fill="x", pady=(6, 0))

        # 2. Сводка результатов
        frame_summary = ttk.LabelFrame(
            self.tab_diff, text=" 2. Сводка сравнения ", padding=8
        )
        frame_summary.pack(fill="x", **pad_opts)

        self.lbl_diff_f1_info = ttk.Label(
            frame_summary, text="Файл 1: 0 значений (0 уникальных)"
        )
        self.lbl_diff_f1_info.pack(anchor="w")

        self.lbl_diff_f2_info = ttk.Label(
            frame_summary, text="Файл 2: 0 значений (0 уникальных)"
        )
        self.lbl_diff_f2_info.pack(anchor="w")

        self.lbl_diff_stats = ttk.Label(
            frame_summary,
            text="Только в Файле 1: 0  |  Только в Файле 2: 0  |  Общих: 0",
            font=("Segoe UI", 9, "bold"),
        )
        self.lbl_diff_stats.pack(anchor="w", pady=(4, 0))

        # 3. Детальный список результатов
        frame_details = ttk.LabelFrame(
            self.tab_diff, text=" 3. Просмотр результатов Diff ", padding=8
        )
        frame_details.pack(fill="both", expand=True, **pad_opts)

        # Переключатель видов Diff
        f_controls = ttk.Frame(frame_details)
        f_controls.pack(fill="x", pady=(0, 6))

        ttk.Label(f_controls, text="Вид отображения:").pack(
            side="left", padx=(0, 4)
        )
        self.var_diff_view_mode = tk.StringVar(value="side_by_side")

        rb_side = ttk.Radiobutton(
            f_controls,
            text="Таблица Diff (Side-by-Side)",
            value="side_by_side",
            variable=self.var_diff_view_mode,
            command=self._on_diff_view_mode_changed,
        )
        rb_side.pack(side="left", padx=4)

        rb_cat = ttk.Radiobutton(
            f_controls,
            text="Фильтр по категориям",
            value="category",
            variable=self.var_diff_view_mode,
            command=self._on_diff_view_mode_changed,
        )
        rb_cat.pack(side="left", padx=4)

        # Выпадающий список категорий (для режима "category")
        self.cmb_diff_cat = ttk.Combobox(
            f_controls,
            state="readonly",
            values=[
                "Только в Файле 1",
                "Только в Файле 2",
                "Общие значения",
            ],
            width=20,
        )
        self.cmb_diff_cat.current(0)
        self.cmb_diff_cat.bind(
            "<<ComboboxSelected>>", lambda e: self._refresh_diff_tree()
        )

        # Чекбокс: скрыть совпадения в Side-by-side
        self.var_hide_equal = tk.BooleanVar(value=False)
        self.chk_hide_equal = ttk.Checkbutton(
            f_controls,
            text="Только различия",
            variable=self.var_hide_equal,
            command=self._refresh_diff_tree,
        )
        self.chk_hide_equal.pack(side="right", padx=4)

        # Контейнер для таблиц Treeview
        self.tree_container = ttk.Frame(frame_details)
        self.tree_container.pack(fill="both", expand=True)

        # Создаем Treeview для Side-by-Side
        columns_side = ("idx", "val1", "status", "val2")
        self.tree_side = ttk.Treeview(
            self.tree_container,
            columns=columns_side,
            show="headings",
            selectmode="browse",
        )
        self.tree_side.heading("idx", text="№")
        self.tree_side.heading("val1", text="Файл 1")
        self.tree_side.heading("status", text="Статус")
        self.tree_side.heading("val2", text="Файл 2")

        self.tree_side.column("idx", width=50, anchor="center")
        self.tree_side.column("val1", width=250, anchor="w")
        self.tree_side.column("status", width=140, anchor="center")
        self.tree_side.column("val2", width=250, anchor="w")

        # Настраиваем цветовые теги для наглядности diff
        self.tree_side.tag_configure("only1", foreground="#c0392b")  # Красный
        self.tree_side.tag_configure("only2", foreground="#d35400")  # Оранжевый
        self.tree_side.tag_configure("equal", foreground="#27ae60")  # Зеленый

        # Создаем Treeview для Category List
        columns_cat = ("idx", "value")
        self.tree_cat = ttk.Treeview(
            self.tree_container,
            columns=columns_cat,
            show="headings",
            selectmode="browse",
        )
        self.tree_cat.heading("idx", text="№")
        self.tree_cat.heading("value", text="Значение")
        self.tree_cat.column("idx", width=60, anchor="center")
        self.tree_cat.column("value", width=620, anchor="w")

        # Общий Scrollbar
        self.sb_diff = ttk.Scrollbar(
            self.tree_container, orient="vertical", command=self._on_diff_scroll
        )
        self.tree_side.configure(yscrollcommand=self.sb_diff.set)
        self.tree_cat.configure(yscrollcommand=self.sb_diff.set)

        # По умолчанию отображаем side_by_side
        self.tree_side.pack(side="left", fill="both", expand=True)
        self.sb_diff.pack(side="right", fill="y")

        # 4. Панель действий внизу
        frame_diff_actions = ttk.Frame(self.tab_diff)
        frame_diff_actions.pack(fill="x", padx=10, pady=(0, 10))

        ttk.Button(
            frame_diff_actions,
            text="📊 Сохранить отчет в Excel (.xlsx)...",
            command=self._export_diff_to_excel,
        ).pack(side="right", padx=3)
        ttk.Button(
            frame_diff_actions,
            text="💾 Экспорт списка в TXT...",
            command=self._export_diff_to_txt,
        ).pack(side="right", padx=3)
        ttk.Button(
            frame_diff_actions,
            text="📋 Скопировать список",
            command=self._copy_diff_to_clipboard,
        ).pack(side="right", padx=3)

        self.diff_results = None

    def _on_diff_scroll(self, *args):
        if self.var_diff_view_mode.get() == "side_by_side":
            self.tree_side.yview(*args)
        else:
            self.tree_cat.yview(*args)

    def _on_diff_view_mode_changed(self):
        mode = self.var_diff_view_mode.get()
        if mode == "side_by_side":
            self.cmb_diff_cat.pack_forget()
            self.chk_hide_equal.pack(side="right", padx=4)
            self.tree_cat.pack_forget()
            self.tree_side.pack(side="left", fill="both", expand=True)
        else:
            self.chk_hide_equal.pack_forget()
            self.cmb_diff_cat.pack(side="left", padx=4)
            self.tree_side.pack_forget()
            self.tree_cat.pack(side="left", fill="both", expand=True)
        self._refresh_diff_tree()

    def _browse_diff_file1(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if path:
            self.ent_diff_file1.delete(0, tk.END)
            self.ent_diff_file1.insert(0, path)

    def _browse_diff_file2(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if path:
            self.ent_diff_file2.delete(0, tk.END)
            self.ent_diff_file2.insert(0, path)

    def _start_column_diff(self):
        f1 = self.ent_diff_file1.get().strip()
        f2 = self.ent_diff_file2.get().strip()

        if not f1 or not os.path.isfile(f1):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректный Файл 1!"
            )
            return
        if not f2 or not os.path.isfile(f2):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректный Файл 2!"
            )
            return

        try:
            col1 = int(self.spn_diff_col1.get())
            col2 = int(self.spn_diff_col2.get())
        except ValueError:
            messagebox.showwarning(
                "Предупреждение", "Номера колонок должны быть числами!"
            )
            return

        config = {
            "file1": f1,
            "file2": f2,
            "col1": col1,
            "col2": col2,
            "has_header1": self.var_diff_hdr1.get(),
            "has_header2": self.var_diff_hdr2.get(),
        }

        self.btn_diff_run.config(state="disabled")

        threading.Thread(
            target=process_column_diff,
            args=(
                config,
                lambda s, res: self.after(0, lambda: self._on_diff_finish(s, res)),
            ),
            daemon=True,
        ).start()

    def _on_diff_finish(self, success, data):
        self.btn_diff_run.config(state="normal")
        if not success:
            messagebox.showerror("Ошибка сравнения", str(data))
            return

        self.diff_results = data

        f1_name = Path(self.ent_diff_file1.get()).name
        f2_name = Path(self.ent_diff_file2.get()).name

        self.tree_side.heading("val1", text=f"Файл 1 ({f1_name})")
        self.tree_side.heading("val2", text=f"Файл 2 ({f2_name})")

        self.lbl_diff_f1_info.config(
            text=f"Файл 1 ({f1_name}, кол. {self.spn_diff_col1.get()}): {data['total_f1_rows']} строк, {data['unique_f1']} уникальных"
        )
        self.lbl_diff_f2_info.config(
            text=f"Файл 2 ({f2_name}, кол. {self.spn_diff_col2.get()}): {data['total_f2_rows']} строк, {data['unique_f2']} уникальных"
        )
        self.lbl_diff_stats.config(
            text=f"Только в Файле 1: {len(data['only_in_1'])}  |  Только в Файле 2: {len(data['only_in_2'])}  |  Общих: {len(data['common'])}"
        )

        self._refresh_diff_tree()

    def _get_current_diff_list(self) -> list[str]:
        if not self.diff_results:
            return []
        mode = self.var_diff_view_mode.get()
        if mode == "side_by_side":
            rows = self.diff_results.get("side_by_side", [])
            hide_eq = self.var_hide_equal.get()
            return [
                f"{r[0]} | {r[1]} [{r[2]}]"
                for r in rows
                if not (hide_eq and r[3] == "equal")
            ]
        else:
            idx = self.cmb_diff_cat.current()
            if idx == 0:
                return self.diff_results.get("only_in_1", [])
            elif idx == 1:
                return self.diff_results.get("only_in_2", [])
            else:
                return self.diff_results.get("common", [])

    def _refresh_diff_tree(self):
        if not self.diff_results:
            return

        mode = self.var_diff_view_mode.get()

        if mode == "side_by_side":
            for item in self.tree_side.get_children():
                self.tree_side.delete(item)

            rows = self.diff_results.get("side_by_side", [])
            hide_eq = self.var_hide_equal.get()

            row_num = 1
            for val1, val2, status, tag in rows:
                if hide_eq and tag == "equal":
                    continue
                self.tree_side.insert(
                    "",
                    tk.END,
                    values=(row_num, val1, status, val2),
                    tags=(tag,),
                )
                row_num += 1

        else:
            for item in self.tree_cat.get_children():
                self.tree_cat.delete(item)

            items = self._get_current_diff_list()
            for i, val in enumerate(items, 1):
                self.tree_cat.insert("", tk.END, values=(i, val))

    def _copy_diff_to_clipboard(self):
        items = self._get_current_diff_list()
        if not items:
            messagebox.showinfo("Инфо", "Список пуст.")
            return
        self.clipboard_clear()
        self.clipboard_append("\n".join(items))
        messagebox.showinfo(
            "Успех", f"Скопировано {len(items)} значений в буфер обмена."
        )

    def _export_diff_to_txt(self):
        items = self._get_current_diff_list()
        if not items:
            messagebox.showinfo("Инфо", "Нет данных для сохранения.")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt", filetypes=[("Text files", "*.txt")]
        )
        if file_path:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(items))
            messagebox.showinfo("Успех", f"Список сохранен в:\n{file_path}")

    def _export_diff_to_excel(self):
        if not self.diff_results:
            messagebox.showinfo("Инфо", "Сначала выполните сравнение колонок.")
            return

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="diff_report.xlsx",
        )
        if not out_path:
            return

        try:
            wb = openpyxl.Workbook()

            # Лист 1: Side-by-Side Diff
            ws_side = wb.active
            ws_side.title = "Side-by-Side Diff"

            f1_name = Path(self.ent_diff_file1.get()).name
            f2_name = Path(self.ent_diff_file2.get()).name

            ws_side.append([f"Файл 1 ({f1_name})", "Статус", f"Файл 2 ({f2_name})"])
            for r in self.diff_results.get("side_by_side", []):
                ws_side.append([r[0], r[2], r[1]])

            ws_side.column_dimensions["A"].width = 35
            ws_side.column_dimensions["B"].width = 22
            ws_side.column_dimensions["C"].width = 35

            # Лист 2: Категории списков
            ws_cat = wb.create_sheet(title="По категориям")
            only1 = self.diff_results.get("only_in_1", [])
            only2 = self.diff_results.get("only_in_2", [])
            comm = self.diff_results.get("common", [])

            ws_cat.append(
                [
                    f"Только в {f1_name} ({len(only1)})",
                    f"Только в {f2_name} ({len(only2)})",
                    f"Общие значения ({len(comm)})",
                ]
            )

            max_len = max(len(only1), len(only2), len(comm), 1)
            for i in range(max_len):
                v1 = only1[i] if i < len(only1) else ""
                v2 = only2[i] if i < len(only2) else ""
                v3 = comm[i] if i < len(comm) else ""
                ws_cat.append([v1, v2, v3])

            ws_cat.column_dimensions["A"].width = 35
            ws_cat.column_dimensions["B"].width = 35
            ws_cat.column_dimensions["C"].width = 35

            wb.save(out_path)
            messagebox.showinfo(
                "Успех", f"Отчет успешно сохранен в:\n{out_path}"
            )
        except Exception as e:
            messagebox.showerror(
                "Ошибка сохранения", f"Не удалось сохранить Excel: {e}"
            )

    # --------------------------------------------------------------------------
    #                     ВКЛАДКА 5: КОНВЕРТЕР XLSX -> PDF
    # --------------------------------------------------------------------------
    def _build_tab_pdf(self):
        pad_opts = {"padx": 10, "pady": 5}

        self._add_tab_description(
            self.tab_pdf,
            "Конвертируйте Excel-файлы из папки (вместе со всеми подпапками) в PDF в один клик. "
            "Выберите, как называть PDF: 1) по имени исходного Excel-файла (report.xlsx -> report.pdf) или 2) по имени папки, где лежит файл "
            "(папка_123/report.xlsx -> папка_123.pdf — удобно, когда в каждой подпапке один файл). "
            "Галочка «Сохранять структуру папок» — создаст в выходной папке такие же подпапки, как в исходной; без неё все PDF лягут в одну папку. "
            "При совпадении имён добавится _1, _2. Приоритет движков: LibreOffice -> Microsoft Excel (если LibreOffice не установлен) -> reportlab (упрощённо).",
        )

        frame_paths = ttk.LabelFrame(
            self.tab_pdf, text=" 1. Папки ", padding=10
        )
        frame_paths.pack(fill="x", **pad_opts)

        # Исходная папка
        f_src = ttk.Frame(frame_paths)
        f_src.pack(fill="x", pady=2)
        ttk.Label(f_src, text="Папка с XLSX (с подпапками):", width=28).pack(side="left")
        self.ent_pdf_src = ttk.Entry(f_src)
        self.ent_pdf_src.pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(f_src, text="Обзор...", command=self._browse_pdf_src).pack(side="left", padx=2)

        # Выходная папка
        f_out = ttk.Frame(frame_paths)
        f_out.pack(fill="x", pady=4)
        ttk.Label(f_out, text="Папка для PDF:", width=28).pack(side="left")
        self.ent_pdf_out = ttk.Entry(f_out)
        self.ent_pdf_out.pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(f_out, text="Обзор...", command=self._browse_pdf_out).pack(side="left", padx=2)

        # Опции именования
        frame_naming = ttk.LabelFrame(
            self.tab_pdf, text=" 2. Как называть PDF файлы ", padding=10
        )
        frame_naming.pack(fill="x", **pad_opts)

        self.var_pdf_naming = tk.StringVar(value="filename")

        ttk.Radiobutton(
            frame_naming,
            text="1 — Имя исходного файла  (report.xlsx  ->  report.pdf)",
            value="filename",
            variable=self.var_pdf_naming,
        ).pack(anchor="w", pady=2)

        ttk.Radiobutton(
            frame_naming,
            text="2 — Имя родительской папки  (папка_123/report.xlsx  ->  папка_123.pdf)",
            value="parent_folder",
            variable=self.var_pdf_naming,
        ).pack(anchor="w", pady=2)

        ttk.Label(
            frame_naming,
            text="При совпадении имён автоматически добавится суффикс _1, _2 ...  •  Недопустимые символы заменяются на _",
            font=("Segoe UI", 8),
            foreground="#666",
        ).pack(anchor="w", pady=(6, 0))

        # Опция сохранения структуры
        self.var_pdf_keep_structure = tk.BooleanVar(value=False)
        chk_keep = ttk.Checkbutton(
            frame_naming,
            text="Сохранять структуру папок  (в выходной папке создавать такие же подпапки, как в исходной)",
            variable=self.var_pdf_keep_structure,
        )
        chk_keep.pack(anchor="w", pady=(10, 2))
        ttk.Label(
            frame_naming,
            text="Если включено:  исходная/2024/отчет.xlsx  ->  выходная/2024/отчет.pdf  (или .../2024.pdf при выборе «имя папки»).\nЕсли выключено — все PDF складываются в одну папку.",
            font=("Segoe UI", 8),
            foreground="#666",
            justify="left",
        ).pack(anchor="w", padx=(26, 0), pady=(0, 2))

        # Выбор движка
        frame_engine = ttk.LabelFrame(self.tab_pdf, text=" 3. Движок конвертации ", padding=10)
        frame_engine.pack(fill="x", **pad_opts)

        self.var_pdf_engine = tk.StringVar(value="auto")
        ttk.Radiobutton(frame_engine, text="Авто (рекомендуется) — LibreOffice -> Excel -> reportlab", value="auto", variable=self.var_pdf_engine).pack(anchor="w", pady=1)
        ttk.Radiobutton(frame_engine, text="Только LibreOffice (точная, сохраняет форматирование)", value="libre", variable=self.var_pdf_engine).pack(anchor="w", pady=1)
        ttk.Radiobutton(frame_engine, text="Только Microsoft Excel (точная, через COM — нужен Excel)", value="excel", variable=self.var_pdf_engine).pack(anchor="w", pady=1)
        ttk.Radiobutton(frame_engine, text="Только reportlab (упрощённая таблица, без Excel/LibreOffice)", value="reportlab", variable=self.var_pdf_engine).pack(anchor="w", pady=1)
        ttk.Radiobutton(frame_engine, text="Только Google Sheets (через Drive API — нужен credentials.json)", value="google", variable=self.var_pdf_engine).pack(anchor="w", pady=1)

        # Инфо о доступных движках
        soffice_path = find_soffice_executable()
        excel_ok = _is_excel_available()
        lines = []
        if soffice_path:
            lines.append(f"✓ LibreOffice: найдено ({soffice_path})")
        else:
            lines.append("✗ LibreOffice: не найдено")
        if excel_ok:
            lines.append("✓ Microsoft Excel: доступен (COM)")
        else:
            if platform.system() == "Windows":
                lines.append("✗ Microsoft Excel: не найден")
            else:
                lines.append("— Microsoft Excel: доступен только на Windows")
        try:
            import reportlab  # noqa: F401

            has_rl = True
        except ImportError:
            has_rl = False
        if has_rl:
            lines.append("✓ reportlab: установлен")
        else:
            lines.append("✗ reportlab: не установлен")
        # Google — проверяем наличие библиотек и credentials
        try:
            import googleapiclient  # noqa: F401

            has_google_lib = True
        except ImportError:
            has_google_lib = False
        cred_default = str(Path(__file__).parent / "credentials.json")
        has_cred = os.path.isfile(cred_default)
        if has_google_lib and has_cred:
            lines.append("✓ Google: готов (библиотеки + credentials.json)")
        elif has_google_lib:
            lines.append("○ Google: библиотеки есть, нет credentials.json")
        else:
            lines.append("✗ Google: нет библиотек (pip install google-api-python-client)")

        if soffice_path:
            auto_hint = "Авто -> LibreOffice"
            fg = "#1a7f37"
        elif excel_ok:
            auto_hint = "Авто -> Excel"
            fg = "#1a7f37"
        elif has_rl:
            auto_hint = "Авто -> reportlab"
            fg = "#b7791f"
        else:
            auto_hint = "Нет доступных движков"
            fg = "#c0392b"
        engine_info = " | ".join(lines) + f"  ->  {auto_hint}."
        ttk.Label(frame_engine, text=engine_info, font=("Segoe UI", 8), foreground=fg, wraplength=780, justify="left", anchor="w").pack(anchor="w", pady=(6, 0))

        # Настройки Google (видны только при выборе Google)
        self.frame_google = ttk.LabelFrame(self.tab_pdf, text=" Настройки Google Sheets ", padding=10)
        # не pack сразу, покажем по выбору движка

        f_cred = ttk.Frame(self.frame_google)
        f_cred.pack(fill="x", pady=2)
        ttk.Label(f_cred, text="credentials.json:", width=18).pack(side="left")
        self.ent_google_creds = ttk.Entry(f_cred)
        self.ent_google_creds.pack(side="left", fill="x", expand=True, padx=5)
        # по умолчанию credentials.json рядом с app.py
        default_cred = str(Path(__file__).parent / "credentials.json")
        self.ent_google_creds.insert(0, default_cred)
        ttk.Button(f_cred, text="Обзор...", command=self._browse_google_creds).pack(side="left", padx=2)

        f_tok = ttk.Frame(self.frame_google)
        f_tok.pack(fill="x", pady=2)
        ttk.Label(f_tok, text="token.json:", width=18).pack(side="left")
        self.ent_google_token = ttk.Entry(f_tok)
        self.ent_google_token.pack(side="left", fill="x", expand=True, padx=5)
        default_tok = str(Path(__file__).parent / "token.json")
        self.ent_google_token.insert(0, default_tok)
        ttk.Button(f_tok, text="Обзор...", command=self._browse_google_token).pack(side="left", padx=2)

        ttk.Label(
            self.frame_google,
            text="1. Создайте проект в Google Cloud Console -> Drive API -> Credentials -> OAuth client ID (Desktop).\n"
            "2. Скачайте credentials.json и укажите путь. При первой конвертации откроется браузер для входа.",
            font=("Segoe UI", 8),
            foreground="#666",
            justify="left",
        ).pack(anchor="w", pady=(6, 0))
        ttk.Button(self.frame_google, text="🔑 Проверить подключение Google", command=self._test_google_connection).pack(anchor="w", pady=(6, 0))

        # Обновляем видимость Google-фрейма при переключении движка
        for rb in frame_engine.winfo_children():
            try:
                rb.configure(command=self._on_pdf_engine_changed)
            except Exception:
                pass
        self._on_pdf_engine_changed()

        # Кнопка запуска
        self.btn_pdf_run = ttk.Button(
            self.tab_pdf,
            text="▶ Конвертировать в PDF",
            command=self._start_pdf_convert,
        )
        self.btn_pdf_run.pack(fill="x", padx=10, pady=6)

        self.pdf_progress = ttk.Progressbar(self.tab_pdf, mode="determinate")
        self.pdf_progress.pack(fill="x", padx=10, pady=2)

        frame_log = ttk.LabelFrame(self.tab_pdf, text=" 4. Журнал ", padding=8)
        frame_log.pack(fill="both", expand=True, **pad_opts)

        self.txt_pdf_log = tk.Text(frame_log, height=12, wrap="word", font=("Consolas", 9))
        self.txt_pdf_log.pack(fill="both", expand=True)

        # Нижняя панель действий
        frame_pdf_actions = ttk.Frame(self.tab_pdf)
        frame_pdf_actions.pack(fill="x", padx=10, pady=(0, 8))

        ttk.Button(
            frame_pdf_actions,
            text="📂 Открыть папку с PDF",
            command=self._open_pdf_output_folder,
        ).pack(side="right", padx=3)
        ttk.Button(
            frame_pdf_actions,
            text="Очистить журнал",
            command=lambda: self.txt_pdf_log.delete(1.0, tk.END),
        ).pack(side="right", padx=3)

    def _on_pdf_engine_changed(self):
        """Показывает/скрывает настройки Google в зависимости от выбранного движка."""
        if not hasattr(self, "frame_google") or not hasattr(self, "var_pdf_engine"):
            return
        eng = self.var_pdf_engine.get()
        if eng == "google":
            # Показать перед кнопкой запуска
            try:
                self.frame_google.pack(fill="x", padx=10, pady=5, before=self.btn_pdf_run)
            except Exception:
                self.frame_google.pack(fill="x", padx=10, pady=5)
        else:
            self.frame_google.pack_forget()

    def _browse_google_creds(self):
        path = filedialog.askopenfilename(title="Выберите credentials.json", filetypes=[("JSON", "*.json"), ("All", "*.*")])
        if path:
            self.ent_google_creds.delete(0, tk.END)
            self.ent_google_creds.insert(0, path)
            # по умолчанию token рядом с credentials
            tok = str(Path(path).parent / "token.json")
            self.ent_google_token.delete(0, tk.END)
            self.ent_google_token.insert(0, tok)

    def _browse_google_token(self):
        path = filedialog.askopenfilename(title="Выберите token.json", filetypes=[("JSON", "*.json"), ("All", "*.*")])
        if path:
            self.ent_google_token.delete(0, tk.END)
            self.ent_google_token.insert(0, path)

    def _test_google_connection(self):
        cred = self.ent_google_creds.get().strip() if hasattr(self, "ent_google_creds") else "credentials.json"
        tok = self.ent_google_token.get().strip() if hasattr(self, "ent_google_token") else "token.json"
        if not cred or not os.path.isfile(cred):
            messagebox.showwarning("Проверка Google", f"Не найден файл credentials.json:\n{cred}")
            return
        try:
            self.txt_pdf_log.insert(tk.END, f"Проверка Google Drive... {cred}\n")
            self.txt_pdf_log.see(tk.END)
            service, creds = _get_google_drive_service(cred, tok)
            # пробный запрос — список файлов (1)
            service.files().list(pageSize=1, fields="files(id,name)").execute()
            messagebox.showinfo("Проверка Google", "Подключение успешно! Токен сохранён.")
            self.txt_pdf_log.insert(tk.END, "Google Drive: подключение успешно\n")
        except Exception as e:
            messagebox.showerror("Проверка Google", f"Ошибка подключения:\n{e}")
            self.txt_pdf_log.insert(tk.END, f"Google ошибка: {e}\n")
        self.txt_pdf_log.see(tk.END)

    def _browse_pdf_src(self):
        path = filedialog.askdirectory(title="Выберите папку с XLSX файлами")
        if path:
            self.ent_pdf_src.delete(0, tk.END)
            self.ent_pdf_src.insert(0, path)
            if not self.ent_pdf_out.get():
                p = Path(path)
                # Предлагаем выходную папку рядом
                self.ent_pdf_out.insert(0, str(p.parent / f"{p.name}_pdf"))

    def _browse_pdf_out(self):
        path = filedialog.askdirectory(title="Выберите папку для сохранения PDF")
        if path:
            self.ent_pdf_out.delete(0, tk.END)
            self.ent_pdf_out.insert(0, path)

    def _open_pdf_output_folder(self):
        folder = self.ent_pdf_out.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("Предупреждение", "Папка с PDF ещё не создана или не выбрана.")
            return
        try:
            system_name = platform.system()
            if system_name == "Windows":
                os.startfile(folder)
            elif system_name == "Darwin":
                subprocess.run(["open", folder], check=True)
            else:
                subprocess.run(["xdg-open", folder], check=True)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть папку: {e}")

    def _start_pdf_convert(self):
        src = self.ent_pdf_src.get().strip()
        out = self.ent_pdf_out.get().strip()
        naming = self.var_pdf_naming.get()
        keep = self.var_pdf_keep_structure.get()
        engine = self.var_pdf_engine.get() if hasattr(self, "var_pdf_engine") else "auto"
        gcreds = self.ent_google_creds.get().strip() if hasattr(self, "ent_google_creds") else str(Path(__file__).parent / "credentials.json")
        gtok = self.ent_google_token.get().strip() if hasattr(self, "ent_google_token") else str(Path(__file__).parent / "token.json")

        if not src or not os.path.isdir(src):
            messagebox.showwarning("Предупреждение", "Укажите корректную исходную папку с XLSX!")
            return
        if not out:
            messagebox.showwarning("Предупреждение", "Укажите папку для сохранения PDF!")
            return

        self.txt_pdf_log.delete(1.0, tk.END)
        self.pdf_progress["value"] = 0
        self.btn_pdf_run.config(state="disabled")

        threading.Thread(
            target=process_xlsx_to_pdf,
            args=(
                src,
                out,
                naming,
                lambda t: self.after(0, lambda: (self.txt_pdf_log.insert(tk.END, t + "\n"), self.txt_pdf_log.see(tk.END))),
                lambda v: self.after(0, lambda: self.pdf_progress.config(value=v)),
                lambda s, m: self.after(0, lambda: self._on_pdf_finish(s, m)),
                keep,
                engine,
                gcreds,
                gtok,
            ),
            daemon=True,
        ).start()

    def _on_pdf_finish(self, success, msg):
        self.btn_pdf_run.config(state="normal")
        if success:
            messagebox.showinfo("Готово", msg)
        else:
            messagebox.showerror("Ошибка", msg)

    # --------------------------------------------------------------------------
    #                     ОБЩИЕ МЕТОДЫ И БИНДИНГИ
    # --------------------------------------------------------------------------
    def _setup_text_widget_clipboard(self, widget: tk.Text):
        """Поддержка горячих клавиш (включая русскую раскладку и Mac) и меню правой кнопки мыши."""

        def paste_action(event=None):
            try:
                clipboard_text = widget.clipboard_get()
                widget.insert(tk.INSERT, clipboard_text)
            except tk.TclError:
                pass
            return "break"

        def copy_action(event=None):
            try:
                selected_text = widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                widget.clipboard_clear()
                widget.clipboard_append(selected_text)
            except tk.TclError:
                pass
            return "break"

        def cut_action(event=None):
            try:
                selected_text = widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                widget.clipboard_clear()
                widget.clipboard_append(selected_text)
                widget.delete(tk.SEL_FIRST, tk.SEL_LAST)
            except tk.TclError:
                pass
            return "break"

        def select_all_action(event=None):
            widget.tag_add(tk.SEL, "1.0", tk.END)
            widget.mark_set(tk.INSERT, "1.0")
            widget.see(tk.INSERT)
            return "break"

        def clear_action():
            widget.delete("1.0", tk.END)

        # Английская раскладка
        widget.bind("<Control-v>", paste_action)
        widget.bind("<Control-V>", paste_action)
        widget.bind("<Control-c>", copy_action)
        widget.bind("<Control-C>", copy_action)
        widget.bind("<Control-x>", cut_action)
        widget.bind("<Control-X>", cut_action)
        widget.bind("<Control-a>", select_all_action)
        widget.bind("<Control-A>", select_all_action)

        # Русская раскладка
        widget.bind("<Control-Key-Cyrillic_em>", paste_action)
        widget.bind("<Control-Key-Cyrillic_EM>", paste_action)
        widget.bind("<Control-Key-Cyrillic_es>", copy_action)
        widget.bind("<Control-Key-Cyrillic_ES>", copy_action)
        widget.bind("<Control-Key-Cyrillic_che>", cut_action)
        widget.bind("<Control-Key-Cyrillic_CHE>", cut_action)
        widget.bind("<Control-Key-Cyrillic_ef>", select_all_action)
        widget.bind("<Control-Key-Cyrillic_EF>", select_all_action)

        # macOS (Command)
        widget.bind("<Command-v>", paste_action)
        widget.bind("<Command-V>", paste_action)
        widget.bind("<Command-c>", copy_action)
        widget.bind("<Command-C>", copy_action)
        widget.bind("<Command-x>", cut_action)
        widget.bind("<Command-X>", cut_action)
        widget.bind("<Command-a>", select_all_action)
        widget.bind("<Command-A>", select_all_action)

        # Контекстное меню по правому клику
        context_menu = tk.Menu(widget, tearoff=0)
        context_menu.add_command(label="Вставить", command=paste_action)
        context_menu.add_command(label="Копировать", command=copy_action)
        context_menu.add_command(label="Вырезать", command=cut_action)
        context_menu.add_separator()
        context_menu.add_command(label="Выделить всё", command=select_all_action)
        context_menu.add_command(label="Очистить", command=clear_action)

        def show_context_menu(event):
            widget.focus_set()
            context_menu.tk_popup(event.x_root, event.y_root)

        widget.bind("<Button-3>", show_context_menu)
        widget.bind("<Button-2>", show_context_menu)

    def _browse_single_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.ent_single_dir.delete(0, tk.END)
            self.ent_single_dir.insert(0, path)

    def _start_single_search(self):
        folder = self.ent_single_dir.get().strip()
        raw_text = self.txt_queries.get(1.0, tk.END).strip()

        if not folder or not os.path.isdir(folder):
            messagebox.showwarning(
                "Предупреждение", "Укажите корректную папку с файлами Excel!"
            )
            return

        raw_items = re.split(r"[\n,]+", raw_text)
        queries = [q.strip() for q in raw_items if q.strip()]

        if not queries:
            messagebox.showwarning(
                "Предупреждение", "Введите хотя бы одну строку для поиска!"
            )
            return

        for item in self.tree_results.get_children():
            self.tree_results.delete(item)

        self.single_progress["value"] = 0
        self.btn_single_run.config(state="disabled")
        self.lbl_single_status.config(
            text=f"Идет поиск по {len(queries)} значениям..."
        )

        threading.Thread(
            target=process_multi_query_search,
            args=(
                folder,
                queries,
                lambda v: self.after(
                    0, lambda: self.single_progress.config(value=v)
                ),
                lambda s, m: self.after(0, lambda: self._on_single_finish(s, m)),
                lambda res_tuple: self.after(
                    0,
                    lambda: self.tree_results.insert(
                        "", tk.END, values=res_tuple
                    ),
                ),
            ),
            daemon=True,
        ).start()

    def _on_single_finish(self, success, msg):
        self.btn_single_run.config(state="normal")
        found_count = len(self.tree_results.get_children())
        self.lbl_single_status.config(
            text=f"Найдено файлов: {found_count}" if success else "Ошибка"
        )
        if success:
            messagebox.showinfo("Результат", msg)
        else:
            messagebox.showerror("Ошибка", msg)

    def _open_selected_file(self, event):
        selected = self.tree_results.selection()
        if not selected:
            return
        item_vals = self.tree_results.item(selected[0], "values")
        if len(item_vals) >= 4:
            fullpath = item_vals[3]
            if os.path.exists(fullpath):
                try:
                    system_name = platform.system()
                    if system_name == "Windows":
                        os.startfile(fullpath)
                    elif system_name == "Darwin":
                        subprocess.run(["open", fullpath], check=True)
                    else:
                        subprocess.run(["xdg-open", fullpath], check=True)
                except Exception as e:
                    messagebox.showerror(
                        "Ошибка", f"Не удалось открыть файл: {e}"
                    )

    def _copy_files_to_folder(self):
        items = self.tree_results.get_children()
        if not items:
            messagebox.showinfo(
                "Инфо", "Список пуст. Сначала выполните поиск."
            )
            return

        dest_dir = filedialog.askdirectory(
            title="Выберите папку для копирования найденных файлов"
        )
        if not dest_dir:
            return

        paths = [
            self.tree_results.item(i, "values")[3]
            for i in items
            if len(self.tree_results.item(i, "values")) >= 4
        ]
        unique_paths = list(dict.fromkeys(paths))

        copied_count = 0
        error_count = 0

        for src_path in unique_paths:
            try:
                if os.path.exists(src_path):
                    shutil.copy2(src_path, dest_dir)
                    copied_count += 1
            except Exception:
                error_count += 1

        msg = f"Успешно скопировано файлов: {copied_count} в:\n{dest_dir}"
        if error_count > 0:
            msg += f"\n\nНе удалось скопировать файлов: {error_count}"
        messagebox.showinfo("Копирование завершено", msg)

    def _copy_results_to_clipboard(self):
        items = self.tree_results.get_children()
        if not items:
            messagebox.showinfo("Инфо", "Список пуст.")
            return
        paths = [
            self.tree_results.item(i, "values")[3]
            for i in items
            if len(self.tree_results.item(i, "values")) >= 4
        ]
        self.clipboard_clear()
        self.clipboard_append("\n".join(paths))
        messagebox.showinfo(
            "Успех", f"Скопировано {len(paths)} путей в буфер обмена."
        )

    def _export_results_to_txt(self):
        items = self.tree_results.get_children()
        if not items:
            messagebox.showinfo("Инфо", "Нет данных для сохранения.")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt", filetypes=[("Text files", "*.txt")]
        )
        if file_path:
            paths = [
                self.tree_results.item(i, "values")[3]
                for i in items
                if len(self.tree_results.item(i, "values")) >= 4
            ]
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("\n".join(paths))
            messagebox.showinfo("Успех", f"Результаты сохранены в:\n{file_path}")


if __name__ == "__main__":
    app = ExcelFinderApp()
    app.mainloop()